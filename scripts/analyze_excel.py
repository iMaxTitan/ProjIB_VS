#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Анализ структуры данных Excel для оптимизации импорта"""

import openpyxl
from collections import defaultdict
from datetime import datetime, timedelta

def excel_to_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=int(value))
    return None

def get_week_number(date):
    date = excel_to_date(date)
    if date is None:
        return None
    return date.isocalendar()[1]

def get_quarter(date):
    date = excel_to_date(date)
    if date is None:
        return None
    return (date.month - 1) // 3 + 1

import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'bdib2025.xlsx')
print(f'Loading: {EXCEL_PATH}', flush=True)
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
print(f'Loaded! Sheets: {wb.sheetnames}', flush=True)

for sheet_name in wb.sheetnames:
    if sheet_name not in wb.sheetnames:
        continue

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    print(f'\n{"="*60}')
    print(f'{sheet_name} ({len(rows)} строк)')
    print('='*60)

    # Анализ по разным группировкам
    processes = defaultdict(int)
    main_tasks = defaultdict(int)
    resources = defaultdict(int)

    # Группировки для недельных планов
    by_process_week = defaultdict(int)           # процесс + неделя
    by_process_week_task = defaultdict(int)      # процесс + неделя + задача (текущий)
    by_process_quarter = defaultdict(int)        # процесс + квартал
    by_resource_week = defaultdict(int)          # сотрудник + неделя
    by_task = defaultdict(int)                   # только основная задача

    for row in rows:
        process = str(row[0]).strip() if row[0] else None
        main_task = str(row[1]).strip() if row[1] else None
        resource = str(row[3]).strip() if row[3] else None
        plan_date = row[6]

        if process: processes[process] += 1
        if main_task: main_tasks[main_task[:100]] += 1
        if resource: resources[resource] += 1

        week = get_week_number(plan_date)
        quarter = get_quarter(plan_date)

        if process and week:
            by_process_week[(process, week)] += 1
        if process and week and main_task:
            by_process_week_task[(process, week, main_task[:100])] += 1
        if process and quarter:
            by_process_quarter[(process, quarter)] += 1
        if resource and week:
            by_resource_week[(resource, week)] += 1
        if main_task:
            by_task[main_task[:100]] += 1

    print(f'\nПроцессов: {len(processes)}')
    for p, c in sorted(processes.items(), key=lambda x: -x[1])[:5]:
        print(f'  {p[:55]}: {c}')

    print(f'\nСотрудников: {len(resources)}')
    for r, c in sorted(resources.items(), key=lambda x: -x[1]):
        print(f'  {r}: {c}')

    print(f'\nУникальных "Основна задача": {len(main_tasks)}')
    print('Топ-10:')
    for t, c in sorted(main_tasks.items(), key=lambda x: -x[1])[:10]:
        print(f'  [{c:4d}] {t[:65]}')

    print(f'\n--- ВАРИАНТЫ ГРУППИРОВКИ НЕДЕЛЬНЫХ ПЛАНОВ ---')
    print(f'Вариант 1: Процесс + Неделя = {len(by_process_week)} планов')
    print(f'  (в среднем {len(rows) // len(by_process_week) if by_process_week else 0} задач на план)')

    print(f'Вариант 2: Процесс + Неделя + Задача = {len(by_process_week_task)} планов (текущий)')
    print(f'  (в среднем {len(rows) // len(by_process_week_task) if by_process_week_task else 0} задач на план)')

    print(f'Вариант 3: Сотрудник + Неделя = {len(by_resource_week)} планов')
    print(f'  (в среднем {len(rows) // len(by_resource_week) if by_resource_week else 0} задач на план)')

    print(f'Вариант 4: Процесс + Квартал = {len(by_process_quarter)} планов')
    print(f'  (в среднем {len(rows) // len(by_process_quarter) if by_process_quarter else 0} задач на план)')

wb.close()
