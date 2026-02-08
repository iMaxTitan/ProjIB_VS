# -*- coding: utf-8 -*-
"""
Скрипт импорта данных из Excel (bdib2025.xlsx) в Supabase

Структура импорта:
1. Годовой план (annual_plans) - один на 2025 год
2. Квартальные планы (quarterly_plans) - по процессам и кварталам
3. Недельные планы (weekly_plans) - по "Основна задача" и неделям
4. Задачи (weekly_tasks) - конкретные задачи с часами
"""

import openpyxl
import json
import urllib.request
import uuid
import time
from datetime import datetime, timedelta
from collections import defaultdict

# Supabase credentials
SUPABASE_URL = 'https://lfsdttsvihyejplmzaaz.supabase.co'
SUPABASE_ANON_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imxmc2R0dHN2aWh5ZWpwbG16YWF6Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDQxODU5MjksImV4cCI6MjA1OTc2MTkyOX0.j4iPdZV0177jCafQVVKiKlNI7KYr0dawut9vzjg0Hpk'
SUPABASE_SERVICE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Imxmc2R0dHN2aWh5ZWpwbG16YWF6Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc0NDE4NTkyOSwiZXhwIjoyMDU5NzYxOTI5fQ.GM1wZ1rivVjrmhHhRGlq_UzlwjpOrrL-kpMjoj_Dd-4'

# Маппинг сотрудников Excel -> user_id в базе
USER_MAPPING = {
    'Барановська': '4c63c960-7d05-4c09-9719-611f89a13c62',
    'Чмух': '3942dbf1-0740-4d44-8f08-23637b98cae2',
    'Лобань Ю.': 'a88667cf-db05-4a52-8f5c-1f16878c3393',
    'Бондаренко': 'ba856d59-1b30-4bcb-8dcd-6821b97b0f1e'
}

# Маппинг компаний Excel -> company_id (нормализуем названия)
COMPANY_MAPPING = {
    'АТБ-Маркет': '805be13b-5cc8-4084-ab0b-3c45ca6e89e6',
    'АТБ-Енерго': 'ffdbad51-be9b-470e-a9bb-1eff1c6596a2',
    'КФ-Квітень': '31d44859-64d5-4201-a36a-65d0e94e9cc8',
    'МФ Фаворит': 'd7129c09-dbb2-4da2-a6c4-f82b3fee298b',
    'ЧП Транс Логістик': '6211ed9d-dc19-4026-ad1e-b049f2e3ee3d',
    'Логістік Юніон': '5c54315a-e43d-45e6-892f-480c2e0e5d84',
    'Рітейл Девелопнент': 'a45d4458-aa01-4106-8202-675521854b21',
    'Корпорація АТБ': '48a6e00b-ef2c-4d3e-bf58-e0d1be17f3c6'
}

# Маппинг процессов Excel -> process_id
PROCESS_MAPPING = {
    'Управління життєвим циклом ІС': '19e0c6a4-0a3a-49c2-a4ac-27fe93447ef6',
    'Управлінська діяльність': 'bf6e6731-27fd-4336-a0af-8d045d50264b',
    'Проєктна діяльність': '17b5fb16-5f2d-46f8-b409-2870aa2d4e9f',
    'Управління подіями ІБ': 'e4f8da3d-e0de-4bd7-8999-73992203820a',
    'Управління правами доступу до ІС та мережевих ресурсів ': 'd131d36c-05ef-4e3c-a5dc-3e52eca89947',
    'Управлінська та організаційна діяльність': '4bbb4e4c-6346-465f-8105-ff2b19043a98',
    'Управління інцидентами ІБ': '85c17213-5805-482e-bac9-1a30de28c852',
    'Управління документацією в сфері ІБ': '21a14ed2-1e15-407a-bae4-6d48a93bc42c'
}

# ID отдела СВК
DEPARTMENT_ID = '9beab000-39d0-4d7a-952d-242cef86d0f0'

# ID пользователя для создания планов (Бондаренко - head отдела)
CREATOR_USER_ID = 'ba856d59-1b30-4bcb-8dcd-6821b97b0f1e'


def supabase_request(endpoint, method='GET', data=None):
    """Выполнить запрос к Supabase REST API"""
    url = f'{SUPABASE_URL}/rest/v1/{endpoint}'
    headers = {
        'apikey': SUPABASE_ANON_KEY,
        'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=representation'
    }

    if data:
        data = json.dumps(data).encode('utf-8')

    req = urllib.request.Request(url, data=data, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            time.sleep(0.1)  # Небольшая пауза между запросами
            return json.loads(response.read().decode('utf-8'))
    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8')
        print(f'Error {e.code}: {error_body}')
        return None
    except urllib.error.URLError as e:
        print(f'Network error: {e}')
        time.sleep(1)  # Подождать и попробовать снова
        try:
            with urllib.request.urlopen(req, timeout=30) as response:
                return json.loads(response.read().decode('utf-8'))
        except:
            return None


def get_week_number(date):
    """Получить номер недели ISO"""
    if isinstance(date, str):
        date = datetime.fromisoformat(date.replace(' 00:00:00', ''))
    return date.isocalendar()[1]


def get_quarter(date):
    """Получить квартал из даты"""
    if isinstance(date, str):
        date = datetime.fromisoformat(date.replace(' 00:00:00', ''))
    return (date.month - 1) // 3 + 1


def get_monday_of_week(year, week_num):
    """Получить понедельник недели по номеру"""
    jan1 = datetime(year, 1, 1)
    # Найти первый понедельник года
    days_to_monday = (7 - jan1.weekday()) % 7
    first_monday = jan1 + timedelta(days=days_to_monday)
    # Если первый понедельник после 4 января, это неделя 1
    if first_monday.day > 4:
        first_monday -= timedelta(days=7)
    # Добавить нужное количество недель
    return first_monday + timedelta(weeks=week_num - 1)


def normalize_company_name(name):
    """Нормализовать название компании для маппинга"""
    if not name:
        return None
    # Убираем лишние пробелы и нормализуем
    name = name.strip()
    # Пробуем найти по частичному совпадению
    for key in COMPANY_MAPPING:
        if key.lower().replace('-', '').replace(' ', '') == name.lower().replace('-', '').replace(' ', ''):
            return COMPANY_MAPPING[key]
    return None


def read_excel_data(filepath):
    """Читаем данные из Excel"""
    print(f'Читаю файл: {filepath}')
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Лист1']

    # Получаем заголовки
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val) if val else f'Col{col}')

    col_indices = {h: i+1 for i, h in enumerate(headers)}

    # Читаем все строки
    rows = []
    for row_num in range(2, ws.max_row + 1):
        row_data = {}
        for header in headers:
            val = ws.cell(row=row_num, column=col_indices[header]).value
            if val is not None:
                row_data[header] = val
        if row_data:
            rows.append(row_data)

    print(f'Прочитано {len(rows)} строк')
    return rows


def create_annual_plan():
    """Создаём годовой план на 2025"""
    print('\n=== Создание годового плана ===')

    # Проверяем, есть ли уже план на 2025
    existing = supabase_request('annual_plans?year=eq.2025')
    if existing:
        print(f'Годовой план на 2025 уже существует: {existing[0]["annual_id"]}')
        return existing[0]['annual_id']

    annual_plan = {
        'year': 2025,
        'goal': 'Річний план підрозділу інформаційної безпеки на 2025 рік',
        'expected_result': 'Виконання всіх запланованих завдань з інформаційної безпеки',
        'status': 'active',
        'user_id': CREATOR_USER_ID,
        'budget': 0
    }

    result = supabase_request('annual_plans', 'POST', annual_plan)
    if result:
        annual_id = result[0]['annual_id']
        print(f'Создан годовой план: {annual_id}')
        return annual_id
    return None


def create_quarterly_plans(annual_plan_id, rows):
    """Создаём квартальные планы по процессам"""
    print('\n=== Создание квартальных планов ===')

    # Группируем по (процесс, квартал)
    quarterly_groups = defaultdict(set)
    for row in rows:
        process_name = row.get('Процесс')
        plan_date = row.get('Планова дата')
        if process_name and plan_date:
            quarter = get_quarter(plan_date)
            quarterly_groups[(process_name, quarter)].add(row.get('Основна задача', ''))

    quarterly_plans = {}  # (process_name, quarter) -> quarterly_id

    for (process_name, quarter), tasks in quarterly_groups.items():
        process_id = PROCESS_MAPPING.get(process_name)
        if not process_id:
            print(f'  Процесс не найден в маппинге: {process_name}')
            continue

        # Проверяем, есть ли уже такой план
        existing = supabase_request(
            f'quarterly_plans?annual_plan_id=eq.{annual_plan_id}&process_id=eq.{process_id}&quarter=eq.{quarter}'
        )

        if existing:
            quarterly_id = existing[0]['quarterly_id']
            print(f'  Q{quarter} {process_name[:30]}... уже существует')
        else:
            # Формируем цель из списка задач
            task_list = list(tasks)[:5]  # Максимум 5 задач в описании
            goal = f'{process_name} - Q{quarter}/2025'
            expected_result = '; '.join(task_list)
            if len(expected_result) > 500:
                expected_result = expected_result[:497] + '...'

            quarterly_plan = {
                'annual_plan_id': annual_plan_id,
                'department_id': DEPARTMENT_ID,
                'process_id': process_id,
                'quarter': quarter,
                'goal': goal,
                'expected_result': expected_result,
                'status': 'active'
            }

            result = supabase_request('quarterly_plans', 'POST', quarterly_plan)
            if result:
                quarterly_id = result[0]['quarterly_id']
                print(f'  Создан: Q{quarter} {process_name[:30]}...')
            else:
                continue

        quarterly_plans[(process_name, quarter)] = quarterly_id

    print(f'Всего квартальных планов: {len(quarterly_plans)}')
    return quarterly_plans


def create_weekly_plans(quarterly_plans, rows):
    """Создаём недельные планы по "Основна задача" и неделям"""
    print('\n=== Создание недельных планов ===')

    # Группируем по (процесс, квартал, основная задача, неделя)
    weekly_groups = defaultdict(lambda: {'rows': [], 'users': set(), 'companies': set(), 'hours': 0})

    for row in rows:
        process_name = row.get('Процесс')
        plan_date = row.get('Планова дата')
        main_task = row.get('Основна задача')
        week_str = row.get('Неделя')  # "1 2025"

        if not all([process_name, plan_date, main_task, week_str]):
            continue

        quarter = get_quarter(plan_date)
        week_num = int(week_str.split()[0])

        key = (process_name, quarter, main_task, week_num)
        weekly_groups[key]['rows'].append(row)

        # Собираем исполнителей
        user = row.get('Ресурси')
        if user and user in USER_MAPPING:
            weekly_groups[key]['users'].add(USER_MAPPING[user])

        # Собираем компании
        company = row.get('T_Inf.Организация')
        company_id = normalize_company_name(company)
        if company_id:
            weekly_groups[key]['companies'].add(company_id)

        # Суммируем плановые часы
        plan_hours = row.get('План ч/г')
        if plan_hours:
            try:
                weekly_groups[key]['hours'] += float(plan_hours)
            except:
                pass

    weekly_plans = {}  # key -> weekly_id
    created_count = 0

    for key, data in weekly_groups.items():
        process_name, quarter, main_task, week_num = key

        # Получаем quarterly_id
        quarterly_id = quarterly_plans.get((process_name, quarter))
        if not quarterly_id:
            continue

        # Дата понедельника недели
        monday = get_monday_of_week(2025, week_num)
        weekly_date = monday.strftime('%Y-%m-%d')

        # Проверяем, есть ли уже такой план
        # Ищем по quarterly_id, дате и похожему описанию
        existing = supabase_request(
            f'weekly_plans?quarterly_id=eq.{quarterly_id}&weekly_date=eq.{weekly_date}'
        )

        # Проверяем, есть ли план с таким же описанием
        weekly_id = None
        if existing:
            for plan in existing:
                if main_task in plan.get('expected_result', ''):
                    weekly_id = plan['weekly_id']
                    break

        if weekly_id:
            print(f'  Неделя {week_num}: {main_task[:40]}... уже существует')
        else:
            weekly_plan = {
                'quarterly_id': quarterly_id,
                'weekly_date': weekly_date,
                'expected_result': main_task,
                'planned_hours': min(data['hours'], 168),  # Максимум 168 часов в неделю
                'status': 'active'
            }

            result = supabase_request('weekly_plans', 'POST', weekly_plan)
            if result:
                weekly_id = result[0]['weekly_id']
                created_count += 1

                # Добавляем исполнителей
                for user_id in data['users']:
                    assignee = {
                        'weekly_plan_id': weekly_id,
                        'user_id': user_id
                    }
                    supabase_request('weekly_plan_assignees', 'POST', assignee)

                # Добавляем компании
                for company_id in data['companies']:
                    company_link = {
                        'weekly_id': weekly_id,
                        'company_id': company_id
                    }
                    supabase_request('weekly_plan_companies', 'POST', company_link)

                if created_count % 50 == 0:
                    print(f'  Создано {created_count} недельных планов...')

        weekly_plans[key] = weekly_id

    print(f'Создано недельных планов: {created_count}')
    return weekly_plans


def create_weekly_tasks(weekly_plans, rows):
    """Создаём задачи для недельных планов"""
    print('\n=== Создание задач ===')

    # Группируем задачи по (weekly_plan_key, конкретная задача, исполнитель)
    task_groups = defaultdict(lambda: {'hours': 0, 'completed_at': None, 'document': '', 'note': ''})

    for row in rows:
        process_name = row.get('Процесс')
        plan_date = row.get('Планова дата')
        main_task = row.get('Основна задача')
        week_str = row.get('Неделя')
        task_name = row.get('Задача')
        user = row.get('Ресурси')

        if not all([process_name, plan_date, main_task, week_str, task_name, user]):
            continue

        quarter = get_quarter(plan_date)
        week_num = int(week_str.split()[0])

        weekly_key = (process_name, quarter, main_task, week_num)
        if weekly_key not in weekly_plans:
            continue

        user_id = USER_MAPPING.get(user)
        if not user_id:
            continue

        task_key = (weekly_key, task_name, user_id)

        # Суммируем фактические часы
        fact_hours = row.get('Факт ч/г')
        if fact_hours:
            try:
                task_groups[task_key]['hours'] += float(fact_hours)
            except:
                pass

        # Берём дату выполнения
        completed = row.get('Дата виконання')
        if completed and not task_groups[task_key]['completed_at']:
            if isinstance(completed, str):
                task_groups[task_key]['completed_at'] = completed.split()[0]
            else:
                task_groups[task_key]['completed_at'] = completed.strftime('%Y-%m-%d')

        # Документ и примечание
        doc = row.get('Документ', '')
        note = row.get('Примітка', '')
        if doc:
            task_groups[task_key]['document'] = str(doc)
        if note:
            task_groups[task_key]['note'] = str(note)

    created_count = 0

    for task_key, data in task_groups.items():
        weekly_key, task_name, user_id = task_key
        weekly_id = weekly_plans.get(weekly_key)

        if not weekly_id:
            continue

        # Формируем описание
        description = task_name
        if data['note']:
            description += f'. {data["note"]}'
        if len(description) > 1000:
            description = description[:997] + '...'

        task = {
            'weekly_plan_id': weekly_id,
            'user_id': user_id,
            'description': description,
            'spent_hours': round(data['hours'], 2),
            'completed_at': data['completed_at'],
            'attachment_url': data['document'] if data['document'] else None
        }

        result = supabase_request('weekly_tasks', 'POST', task)
        if result:
            created_count += 1
            if created_count % 100 == 0:
                print(f'  Создано {created_count} задач...')

    print(f'Создано задач: {created_count}')
    return created_count


def main():
    """Основная функция импорта"""
    print('=' * 60)
    print('ИМПОРТ ДАННЫХ ИЗ EXCEL В SUPABASE')
    print('=' * 60)

    # Читаем данные из Excel
    rows = read_excel_data('c:/Proj/ProjIB_VS/bdib2025.xlsx')

    # Создаём годовой план
    annual_plan_id = create_annual_plan()
    if not annual_plan_id:
        print('Ошибка создания годового плана!')
        return

    # Создаём квартальные планы
    quarterly_plans = create_quarterly_plans(annual_plan_id, rows)

    # Создаём недельные планы
    weekly_plans = create_weekly_plans(quarterly_plans, rows)

    # Создаём задачи
    tasks_count = create_weekly_tasks(weekly_plans, rows)

    print('\n' + '=' * 60)
    print('ИМПОРТ ЗАВЕРШЁН')
    print(f'  Годовой план: 1')
    print(f'  Квартальных планов: {len(quarterly_plans)}')
    print(f'  Недельных планов: {len(weekly_plans)}')
    print(f'  Задач: {tasks_count}')
    print('=' * 60)


if __name__ == '__main__':
    main()
