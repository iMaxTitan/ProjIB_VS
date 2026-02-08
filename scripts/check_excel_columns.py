#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'bdib2025.xlsx')
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

ws = wb.active
print(f'Sheet: {ws.title}')

# Заголовки (первая строка)
headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
print('\nКолонки:')
for i, h in enumerate(headers):
    print(f'  [{i}] {h}')

# Примеры первых 5 строк данных
print('\nПримеры данных (5 строк):')
for row in list(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
    print(f'  {row[:10]}...')  # первые 10 колонок

# Проверим даты
print('\nПроверка дат (колонка 6 - Планова дата):')
for row in list(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
    plan_date = row[6] if len(row) > 6 else None
    print(f'  type={type(plan_date).__name__}, value={plan_date}')

wb.close()
